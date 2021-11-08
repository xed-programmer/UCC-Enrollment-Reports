import pyodbc
from openpyxl import load_workbook
import datetime


def create_report(data):
    filename = 'ENROLLMENT-MATRIX.xlsx'
    date = datetime.datetime.now().strftime('%b-%d-%Y-')
    header_date = datetime.datetime.now().strftime('%B, %d %Y ')

    wb = load_workbook('src/' + filename)
    ws = wb.active
    # set the header
    ws['A3'].value += " " + header_date

    cell_sub = 'A'
    cell_start_num = 7
    tot_sub = 36

    for i in range(cell_start_num, tot_sub + 1):
        cell = 'A'
        for d in data:
            if d[0] == ws[cell_sub + str(i)].value:
                n = ord(cell)
                n += 2
                cell = chr(n)
                ws[cell + str(i)] = d[3]

    # Save file
    wb.save('src/' + date + filename)


try:
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=src\Database.accdb;"
    conn = pyodbc.connect(con_string)
    print('Connected to database')

    cur = conn.cursor()
    cur.execute('SELECT COURSE, YEAR1, SEX, COUNT(SEX) FROM student WHERE SEX <> null GROUP BY COURSE, YEAR1, SEX')
    # cur.execute('SELECT COURSE FROM student GROUP BY COURSE')

    row = cur.fetchall()
    create_report(row)

    # Check for null value in SEX Column
    cur.execute('SELECT COURSE, YEAR1, COUNT(STUDE_NO) FROM student WHERE SEX IS NULL GROUP BY COURSE, YEAR1, SEX')
    row = cur.fetchall()
    if len(row) > 0:
        print(f'There are {len(row)} records that is not include in report because they dont have value in SEX Column. Please '
              f'check the database')
        for r in row:
            print(r)

    print('>> Report successfully created!')
    input('>> press any key to exit...')
    exit()

except pyodbc.Error as e:
    print("Error in connection", e)

