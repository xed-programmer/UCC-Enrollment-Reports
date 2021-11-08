import pyodbc
import openpyxl as xl
import datetime, os

date = datetime.datetime.now().strftime('%b-%d-%Y-')
header_date = datetime.datetime.now().strftime('%B, %d %Y ')
filename = "Form-3-ENROLLMENT-LIST-FORMAT.xlsx"
newfilename = date + "Form-3-ENROLLMENT-LIST-FORMAT.xlsx"


def create_new_sheet(subject, year):
    file = 'src/'
    if os.path.exists(file + newfilename):
        file += newfilename
    else:
        file += filename
    wb1 = xl.load_workbook(file)
    ws1 = wb1.worksheets[0]

    # opening the destination sheet
    print('Creating new sheet')
    ws2 = wb1.copy_worksheet(ws1)
    ws2.title = subject + year

    # Save file
    wb1.save('src/' + newfilename)


def create_report(data):
    create_new_sheet(data[0][0], data[0][1])
    file = 'src/'
    if os.path.exists(file + newfilename):
        file += newfilename
    else:
        file += filename
    wb = xl.load_workbook(file)
    ws = wb.worksheets[len(wb.sheetnames) - 1]
    print('Current Sheet: ' + ws.title)

    # set the header
    ws['C5'].value = data[0][0]
    ws['C6'].value = data[0][1] + ' Year'

    current_course = data[0][0]
    current_year = data[0][1]

    cell_start_row = 10
    cell_start_col = 1
    max_col = 19

    i = cell_start_row
    count = 1
    for d in data:
        if d[0] == current_course and d[1] == current_year:
            for j in range(cell_start_col, max_col):
                if j == 1:
                    ws.cell(i, j).value = count
                    count += 1
                elif j >= 8 and j <= 17:
                    ws.cell(i, j).value = d[j + 20] + " - " + d[j]
                elif j != max_col - 1:
                    ws.cell(i, j).value = d[j]
                else:
                    tot_unit = 0
                    # add each unit
                    for k in range(10):
                        tot_unit += d[j + k]
                    ws.cell(i, j).value = tot_unit
            print('Data added')
            data = data[1:]
        else:

            ws.cell(i+2, 1).value = 'PREPARED BY: '
            ws.cell(i + 2, 7).value = 'CERTIFIED CORRECT BY: '
            ws.cell(i + 4, 7).value = 'NAME OF REGISTRAR'
            ws.cell(i + 5, 7).value = 'POSITION'
            if len(data) > 0:
                # Save file
                wb.save('src/' + newfilename)
                create_report(data)
        i += 1
    if len(data) > 0:
        # Save file
        wb.save('src/' + newfilename)
        create_report(data)
    else:
        # Save file
        wb.save('src/' + newfilename)
        print('Done Generating Reports...')
        input('Press enter key to exit...')
        exit()


try:
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=src\Database.accdb;"
    conn = pyodbc.connect(con_string)
    print('Connected to database')

    cur = conn.cursor()
    cur.execute("SELECT COURSE, YEAR1, STUDNOLOC, FNAME, MNAME, SNAME, SEX, STATUS, "
                "TIME1, TIME2, TIME3, TIME4, TIME5, TIME6, TIME7, TIME8, TIME9, TIME10, "
                "UNIT1, UNIT2, UNIT3, UNIT4, UNIT5, UNIT6, UNIT7, UNIT8, UNIT9, UNIT10, "
                "SUBJECT1, SUBJECT2, SUBJECT3, SUBJECT4, SUBJECT5, SUBJECT6, SUBJECT7, SUBJECT8, SUBJECT9, SUBJECT10 "
                "FROM student ORDER BY COURSE, YEAR1, FNAME")

    row = cur.fetchall()
    create_report(list(row))

except pyodbc.Error as e:
    print("Error in connection", e)
