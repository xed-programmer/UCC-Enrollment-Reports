from openpyxl import load_workbook
import datetime, math, os


class EnrollmentReport:

    date = datetime.datetime.now().strftime('%b-%d-%Y-')
    header_date = datetime.datetime.now().strftime('%B, %d %Y ')

    def __init__(self, config):
        self.config = config

    def create_report_by_section(self, data):
        date = datetime.datetime.now().strftime('%b-%d-%Y-')
        filename = self.config['forms'][0]
        header_date = datetime.datetime.now().strftime('%B, %d %Y ').upper()

        wb = load_workbook('src/' + filename)
        ws = wb.active
        # set the header
        ws['A3'].value = self.config['semester'] + " " + ws['A3'].value + " " + header_date

        row_start_num = 7
        row_section = 6
        tot_sub = 36

        for i in range(row_start_num, tot_sub + 1):
            col_num = 3
            prev_year = ''
            for d in data:
                if d[0] == ws.cell(i, 1).value:
                    prev_year = d[0]
                    while d[2] != ws.cell(row_section, col_num).value:
                        col_num += 2
                    ws.cell(i, col_num).value = d[3]
                    col_num += 2
                elif prev_year != '':
                    break

        # Save file
        wb.save('src/' + date + filename)
        return True

    def create_report_by_sex(self, data):
        filename = self.config['forms'][1]
        date = datetime.datetime.now().strftime('%b-%d-%Y-')
        header_date = datetime.datetime.now().strftime('%B, %d %Y ').upper()

        wb = load_workbook('src/' + filename)
        ws = wb.active
        # set the header
        ws['A3'].value = self.config['semester'] + ws['A3'].value + " " + header_date

        row_start_num = 7
        row_sex = 6
        tot_sub = 36

        for i in range(row_start_num, tot_sub + 1):
            col_num = 3
            col_year = 3
            prev_year = ''
            for d in data:
                if d[0] == ws.cell(i, 1).value:
                    prev_year = d[0]
                    while d[1][0] != str(math.floor(col_year/4)+1):
                        col_year += 4

                    col_num = col_year
                    while d[2] != ws.cell(row_sex, col_num).value:
                        col_num += 2
                    ws.cell(i, col_num).value = d[3]
                    col_num += 2
                elif prev_year != '':
                    break

        # Save file
        wb.save('src/' + date + filename)
        return True

    def __create_new_sheet(self, subject, year):
        file = 'src/'
        filename = self.config['forms'][2]
        newfilename = self.date + filename
        if os.path.exists(file + newfilename):
            file += newfilename
        else:
            file += filename
        wb1 = load_workbook(file)
        ws1 = wb1.worksheets[0]

        # opening the destination sheet
        print('Creating new sheet')
        ws2 = wb1.copy_worksheet(ws1)
        ws2.title = subject + year

        # Save file
        wb1.save('src/' + newfilename)

    def create_student_list(self, data):
        filename = self.config['forms'][2]
        newfilename = self.date + filename

        self.__create_new_sheet(data[0][0], data[0][1])
        file = 'src/'
        if os.path.exists(file + newfilename):
            file += newfilename
        else:
            file += filename
        wb = load_workbook(file)
        ws = wb.worksheets[len(wb.sheetnames) - 1]
        print('Current Sheet: ' + ws.title)

        # set the header
        ws['C5'].value = data[0][0]
        ws['C6'].value = data[0][1] + ' Year'

        current_course = data[0][0]
        current_year = data[0][1]

        cell_start_row = 10
        cell_start_col = 1
        max_col = 19

        i = cell_start_row
        count = 1
        for d in data:
            if d[0] == current_course and d[1] == current_year:
                for j in range(cell_start_col, max_col):
                    if j == 1:
                        ws.cell(i, j).value = count
                        count += 1
                    elif j >= 8 and j <= 17:
                        sub = d[j] if d[j] is not None else "*"
                        sub_code = d[j + 20] if d[j + 20] is not None else "*"
                        ws.cell(i, j).value = sub_code + " - " + sub
                    elif j != max_col - 1:
                        ws.cell(i, j).value = d[j]
                    else:
                        tot_unit = 0
                        # add each unit
                        for k in range(10):
                            tot_unit += d[j + k] if d[j + k] is not None else 0
                        ws.cell(i, j).value = tot_unit
                data = data[1:]
            else:

                ws.cell(i + 2, 1).value = 'PREPARED BY: '
                ws.cell(i + 2, 7).value = 'CERTIFIED CORRECT BY: '
                ws.cell(i + 4, 7).value = 'NAME OF REGISTRAR'
                ws.cell(i + 5, 7).value = 'POSITION'
                if len(data) > 0:
                    # Save file
                    wb.save('src/' + newfilename)
                    self.create_student_list(data)
            i += 1
        if len(data) > 0:
            ws.cell(i + 2, 1).value = 'PREPARED BY: '
            ws.cell(i + 2, 7).value = 'CERTIFIED CORRECT BY: '
            ws.cell(i + 4, 7).value = 'NAME OF REGISTRAR'
            ws.cell(i + 5, 7).value = 'POSITION'
            # Save file
            wb.save('src/' + newfilename)
            self.create_student_list(data)
        else:
            ws.cell(i + 2, 1).value = 'PREPARED BY: '
            ws.cell(i + 2, 7).value = 'CERTIFIED CORRECT BY: '
            ws.cell(i + 4, 7).value = 'NAME OF REGISTRAR'
            ws.cell(i + 5, 7).value = 'POSITION'
            # Save file
            wb.save('src/' + newfilename)
            print('>> Report successfully created!')
            input('>> press enter key to exit...')
            exit()
